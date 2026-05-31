# -*- coding: utf-8 -*-
"""
Mini RAG - Model Registry + STAMM Paper + Evaluación LLMs para Bioprocesos
Pipeline: Ingesta YAML + PDF + DOCX + XLSX -> ChromaDB -> Retrieval -> Ollama -> Gradio/Terminal

Modos de uso:
  python app.py              -> Qwen 3.5 9B + Gradio (UI web)
  python app.py --terminal   -> Qwen 3.5 4B + Terminal (ligero)
  python app.py --reingest   -> Forzar re-ingesta de datos
"""

import os
import sys
import json
import yaml
import requests
import chromadb
from collections import Counter

# ============================================================
# CONFIGURACIÓN
# ============================================================
OLLAMA_BASE = "http://localhost:11434"
EMBED_MODEL = "nomic-embed-text-v2-moe"
TOP_K = 5

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECTS_DIR = os.path.join(SCRIPT_DIR, "data", "projects")
PDF_DIR = os.path.join(SCRIPT_DIR, "data")
CHROMA_DIR = os.path.join(SCRIPT_DIR, "chroma_db")

SYSTEM_PROMPT = """Eres un asistente experto en el registro de modelos de machine learning para bioprocesos.
Respondes preguntas sobre:
- Los modelos registrados: configuraciones, inputs, outputs, hiperparámetros y proyectos
- El paper STAMM (Soft sensor moniToring and mAintenance framework for Machine learning Models)
- La evaluación comparativa de LLMs (Gemini, Meta AI, GPT-5.5) sobre preguntas del RAG
- El manual de evaluación del RAG y su rúbrica de calificación
- Qué es un RAG y cómo funcionan los modelos de lenguaje grandes (LLMs)
Usa SOLO la información del contexto proporcionado para responder. Si no encuentras la respuesta en el contexto, dilo claramente.
Responde de forma concisa y precisa. Puedes responder en español o inglés según la pregunta del usuario."""

QUESTION_TEMPLATE = """Utiliza los siguientes fragmentos de contexto para responder la pregunta al final.
Si no sabes la respuesta, di que no lo sabes.
No menciones que te he proporcionado fragmentos, simula que ya tenías esta información en tu conocimiento y responde como en una conversación natural.

{context}

Pregunta: {question}
Respuesta útil:"""

HISTORY_TEMPLATE = """Dada la siguiente conversación y la pregunta, expresa de otro modo la pregunta para que todo sea una sola pregunta en general.

Historial:
{chat_history}
Siguiente pregunta: {question}
Pregunta general:"""


# ============================================================
# EMBEDDINGS
# ============================================================
def ollama_embed(texts):
    """Genera embeddings para una lista de textos."""
    embeddings = []
    for i, text in enumerate(texts):
        resp = requests.post(
            f"{OLLAMA_BASE}/api/embed",
            json={"model": EMBED_MODEL, "input": text},
        )
        resp.raise_for_status()
        embeddings.append(resp.json()["embeddings"][0])
        print(f"  Embedding {i+1}/{len(texts)}", end="\r")
    print(f"  {len(embeddings)} embeddings generados." + " " * 20)
    return embeddings


def ollama_embed_single(text):
    """Genera embedding para un solo texto."""
    resp = requests.post(
        f"{OLLAMA_BASE}/api/embed",
        json={"model": EMBED_MODEL, "input": text},
    )
    resp.raise_for_status()
    return resp.json()["embeddings"][0]


# ============================================================
# CHUNKING DE PDFs
# ============================================================
def pdf_to_chunks(pdf_path, chunk_size=200, overlap=50):
    """Convierte un PDF en chunks de texto por palabras con sliding window."""
    from PyPDF2 import PdfReader

    reader = PdfReader(pdf_path)
    filename = os.path.basename(pdf_path)

    # Extraer todo el texto
    full_text = ""
    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            full_text += page_text + " "

    # Partir en chunks por palabras
    words = full_text.split()
    chunks = []
    start = 0
    chunk_idx = 0

    while start < len(words):
        end = start + chunk_size
        chunk_text = " ".join(words[start:end])

        chunks.append({
            "id": f"pdf_{filename}_{chunk_idx}",
            "text": chunk_text,
            "source": filename,
            "doc_type": "pdf_paper",
        })

        chunk_idx += 1
        start += chunk_size - overlap

    return chunks


# ============================================================
# CHUNKING DE DOCX
# ============================================================
def docx_to_chunks(docx_path, chunk_size=150, overlap=30):
    """Convierte un DOCX en chunks de texto por palabras."""
    from docx import Document

    doc = Document(docx_path)
    filename = os.path.basename(docx_path)

    # Extraer todo el texto párrafo por párrafo
    full_text = " ".join(
        para.text.strip()
        for para in doc.paragraphs
        if para.text.strip()
    )

    words = full_text.split()
    chunks = []
    start = 0
    chunk_idx = 0

    while start < len(words):
        end = start + chunk_size
        chunk_text = " ".join(words[start:end])
        chunks.append({
            "id": f"docx_{filename}_{chunk_idx}",
            "text": chunk_text,
            "source": filename,
            "doc_type": "docx_manual",
        })
        chunk_idx += 1
        start += chunk_size - overlap

    return chunks


# ============================================================
# CHUNKING DE XLSX
# ============================================================
def xlsx_to_chunks(xlsx_path):
    """Convierte un XLSX de evaluación en chunks por pregunta."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    filename = os.path.basename(xlsx_path)
    chunks = []

    # --- Hoja "Respuesta": preguntas + respuestas de cada LLM ---
    if "Respuesta" in wb.sheetnames:
        ws = wb["Respuesta"]
        rows = list(ws.iter_rows(values_only=True))
        # Saltar encabezado (fila 0)
        for row in rows[1:]:
            if row[0] is None or row[2] is None:
                continue
            num = row[0]
            tipo = row[1] or "Normal"
            pregunta = row[2]
            resp_gemini = row[3] or "Sin respuesta"
            resp_meta = row[4] or "Sin respuesta"
            resp_gpt = row[5] or "Sin respuesta"

            text = (
                f"Evaluación RAG - Pregunta {num} ({tipo}): {pregunta}\n\n"
                f"Respuesta Google Gemini 2.5 Flash:\n{resp_gemini}\n\n"
                f"Respuesta Meta AI:\n{resp_meta}\n\n"
                f"Respuesta OpenAI GPT-5.5 Instant:\n{resp_gpt}"
            )
            chunks.append({
                "id": f"xlsx_respuesta_{num}",
                "text": text,
                "source": filename,
                "doc_type": "xlsx_evaluacion",
            })

    # --- Hoja "Calificacion": resumen de puntajes ---
    if "Calificacion " in wb.sheetnames or "Calificacion" in wb.sheetnames:
        sheet_name = "Calificacion " if "Calificacion " in wb.sheetnames else "Calificacion"
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Buscar filas con puntajes totales (contienen "Google", "Meta", "OpenAI")
        resumen_text = "Resultados finales de la evaluación comparativa de LLMs sobre el RAG del Model Registry:\n"
        for row in rows:
            for i, cell in enumerate(row):
                if cell in ("Google (Gemini 2.5 Flash)", "Meta AI online", "OpenAI (GPT-5.5 Instant)"):
                    modelo = cell
                    puntaje = row[i+1] if i+1 < len(row) else None
                    maximo = row[i+2] if i+2 < len(row) else None
                    porcentaje = row[i+3] if i+3 < len(row) else None
                    if puntaje and maximo:
                        resumen_text += f"  - {modelo}: {puntaje}/{maximo} ({porcentaje})\n"

        chunks.append({
            "id": "xlsx_resumen_puntajes",
            "text": resumen_text,
            "source": filename,
            "doc_type": "xlsx_evaluacion",
        })

    # --- Hoja "Puntajes por pregunta": rúbrica ---
    if "Puntajes por pregunta" in wb.sheetnames:
        ws = wb["Puntajes por pregunta"]
        rubrica_text = "Rúbrica de evaluación del RAG (criterios de calificación por respuesta):\n"
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                line = " ".join(str(c) for c in row if c is not None)
                if line.strip():
                    rubrica_text += f"  {line}\n"
        chunks.append({
            "id": "xlsx_rubrica",
            "text": rubrica_text,
            "source": filename,
            "doc_type": "xlsx_evaluacion",
        })

    return chunks


# ============================================================
# CHUNKING DE YAMLS
# ============================================================
def yaml_to_chunks(yaml_path):
    """Convierte un YAML en chunks de texto."""
    with open(yaml_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    chunks = []
    filename = os.path.basename(yaml_path)

    # --- project_info.yaml ---
    if "project_ID" in data:
        text = f"Proyecto: {data.get('project_name', 'N/A')}\n"
        text += f"ID: {data.get('project_ID', 'N/A')}\n"
        text += f"Descripción: {data.get('description', 'N/A')}\n"
        text += f"Coordinador: {data.get('coordinator', 'N/A')}\n"
        text += f"Periodo: {data.get('start_date', '')} - {data.get('end_date', '')}\n"

        if "references" in data:
            refs = "\n".join(r.get("apa", "") for r in data["references"])
            text += f"Referencias:\n{refs}\n"

        if "variables" in data:
            var_lines = []
            for v in data["variables"]:
                var_lines.append(
                    f"  - {v['renamed_variable']} ({v['units']}): {v['description']}"
                )
            text += f"Variables del proceso:\n" + "\n".join(var_lines)

        chunks.append({
            "id": data["project_ID"],
            "text": text,
            "source": filename,
            "doc_type": "project_info",
        })
        return chunks

    # --- config de modelo ---
    cfg = data.get("ml_model_configuration", {})
    ident = cfg.get("model_identification", {})
    desc = cfg.get("model_description", {})
    training = cfg.get("training_information", {})
    inputs = cfg.get("inputs", {})
    outputs = cfg.get("outputs", {})

    model_id = ident.get("ID", "unknown")

    # Chunk 1: Overview
    text1 = f"Modelo: {ident.get('name', 'N/A')}\n"
    text1 += f"ID: {model_id}\n"
    text1 += f"Versión: {ident.get('version', 'N/A')}\n"
    text1 += f"Autor: {ident.get('author', 'N/A')}\n"
    text1 += f"DOI: {ident.get('doi', 'N/A')}\n"
    text1 += f"Fecha de creación: {ident.get('creation_date', 'N/A')}\n"
    text1 += f"Estado: {ident.get('status', 'N/A')} - {ident.get('status_description', '')}\n"
    text1 += f"Tipo: {desc.get('model_type', 'N/A')}\n"
    text1 += f"Learner: {desc.get('learner', 'N/A')}\n"
    text1 += f"Nombre completo: {desc.get('model_name', 'N/A')}\n"
    text1 += f"Descripción: {desc.get('description', 'N/A')}\n"

    langs = desc.get("language", [])
    if langs:
        lang_str = ", ".join(
            f"{l.get('name', '')} {l.get('version', '')}".strip() for l in langs
        )
        text1 += f"Lenguaje: {lang_str}\n"

    pkgs = desc.get("packages", [])
    if pkgs:
        pkg_parts = []
        for p in pkgs:
            if isinstance(p, dict):
                name = p.get("package", "")
                cls = p.get("class", "")
                ver = p.get("version", "")
                label = f"{name}.{cls}" if cls else name
                if ver:
                    label += f" v{ver}"
                pkg_parts.append(label)
            else:
                pkg_parts.append(str(p))
        text1 += f"Paquetes: {', '.join(pkg_parts)}\n"

    chunks.append({
        "id": f"{model_id}_overview",
        "text": text1,
        "source": filename,
        "doc_type": "model_overview",
    })

    # Chunk 2: Training
    if training:
        text2 = f"Entrenamiento del modelo {ident.get('name', '')} (ID: {model_id}):\n"
        text2 += f"Instancias: {training.get('number_of_instances', 'N/A')}\n"
        text2 += f"Validación: {training.get('validation', 'N/A')}\n"

        hyper = training.get("hyperparameters", {})
        if hyper:
            hp_lines = [f"  - {k}: {v}" for k, v in hyper.items()]
            text2 += f"Hiperparámetros:\n" + "\n".join(hp_lines) + "\n"

        chunks.append({
            "id": f"{model_id}_training",
            "text": text2,
            "source": filename,
            "doc_type": "model_training",
        })

    # Chunk 3: Inputs
    features = inputs.get("features", [])
    if features:
        text3 = f"Inputs del modelo {ident.get('name', '')} (ID: {model_id}):\n"
        for feat in features:
            text3 += (
                f"  - {feat['name']} ({feat.get('units', 'N/A')}): "
                f"{feat.get('description', '')} "
                f"[tipo: {feat.get('type', 'N/A')}, "
                f"rango: {feat.get('expected_range', {}).get('min', '?')}"
                f"-{feat.get('expected_range', {}).get('max', '?')}]\n"
            )
        chunks.append({
            "id": f"{model_id}_inputs",
            "text": text3,
            "source": filename,
            "doc_type": "model_inputs",
        })

    # Chunk 4: Outputs
    out_info = outputs.get("information", [])
    if out_info:
        text4 = f"Outputs del modelo {ident.get('name', '')} (ID: {model_id}):\n"
        for out in out_info:
            text4 += (
                f"  - {out['name']} ({out.get('units', 'N/A')}): "
                f"{out.get('description', '')} "
                f"[rango: {out.get('expected_range', {}).get('min', '?')}"
                f"-{out.get('expected_range', {}).get('max', '?')}]\n"
            )
        chunks.append({
            "id": f"{model_id}_outputs",
            "text": text4,
            "source": filename,
            "doc_type": "model_outputs",
        })

    return chunks


# ============================================================
# INGESTA
# ============================================================
def ingest():
    """Recorre proyectos + PDFs, parsea todo, genera embeddings, guarda en ChromaDB."""
    all_chunks = []

    # --- YAMLs ---
    print("  Procesando YAMLs...")
    for project_folder in os.listdir(PROJECTS_DIR):
        project_path = os.path.join(PROJECTS_DIR, project_folder)
        if not os.path.isdir(project_path):
            continue

        info_file = os.path.join(project_path, "project_info.yaml")
        if os.path.exists(info_file):
            all_chunks.extend(yaml_to_chunks(info_file))

        configs_dir = os.path.join(project_path, "configs")
        if os.path.exists(configs_dir):
            for cfg_file in sorted(os.listdir(configs_dir)):
                if cfg_file.endswith(".yaml"):
                    all_chunks.extend(
                        yaml_to_chunks(os.path.join(configs_dir, cfg_file))
                    )

    # --- PDFs ---
    print("  Procesando PDFs...")
    for filename in os.listdir(PDF_DIR):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(PDF_DIR, filename)
            pdf_chunks = pdf_to_chunks(pdf_path, chunk_size=200, overlap=50)
            all_chunks.extend(pdf_chunks)
            print(f"    {filename}: {len(pdf_chunks)} chunks")

    # --- DOCXs ---
    print("  Procesando DOCXs...")
    for filename in os.listdir(PDF_DIR):
        if filename.endswith(".docx"):
            docx_path = os.path.join(PDF_DIR, filename)
            docx_chunks = docx_to_chunks(docx_path, chunk_size=150, overlap=30)
            all_chunks.extend(docx_chunks)
            print(f"    {filename}: {len(docx_chunks)} chunks")

    # --- XLSXs ---
    print("  Procesando XLSXs...")
    for filename in os.listdir(PDF_DIR):
        if filename.endswith(".xlsx"):
            xlsx_path = os.path.join(PDF_DIR, filename)
            xlsx_chunks = xlsx_to_chunks(xlsx_path)
            all_chunks.extend(xlsx_chunks)
            print(f"    {filename}: {len(xlsx_chunks)} chunks")

    if not all_chunks:
        print("No se encontraron documentos.")
        return None

    print(f"\n  {len(all_chunks)} chunks totales:")
    for doc_type, count in Counter(c["doc_type"] for c in all_chunks).items():
        print(f"    {doc_type}: {count}")

    print(f"\n  Generando embeddings...")
    texts = [c["text"] for c in all_chunks]
    embeddings = ollama_embed(texts)

    client = chromadb.PersistentClient(path=CHROMA_DIR)
    try:
        client.delete_collection("model_registry")
    except Exception:
        pass

    collection = client.create_collection(
        name="model_registry",
        metadata={"hnsw:space": "cosine"},
    )

    collection.add(
        ids=[c["id"] for c in all_chunks],
        documents=texts,
        embeddings=embeddings,
        metadatas=[
            {"source": c["source"], "doc_type": c["doc_type"]}
            for c in all_chunks
        ],
    )

    print(f"\n  Ingesta completa: {collection.count()} chunks en ChromaDB.")
    return collection


# ============================================================
# LLM
# ============================================================
class OllamaLLM:
    """LLM usando Ollama API."""

    def __init__(self, model):
        self.model = model

    def completion(self, messages):
        """Genera respuesta con streaming visible en terminal."""
        resp = requests.post(
            f"{OLLAMA_BASE}/api/chat",
            json={
                "model": self.model,
                "messages": messages,
                "stream": True,
                "options": {"num_predict": 300},
                "think": False,
            },
            stream=True,
        )
        resp.raise_for_status()
        full = ""
        print("\nAsistente > ", end="", flush=True)
        for line in resp.iter_lines():
            if line:
                data = json.loads(line)
                token = data.get("message", {}).get("content", "")
                if token:
                    print(token, end="", flush=True)
                    full += token
        print()
        return full

    def completion_stream_print(self, messages):
        """Genera respuesta con streaming visual en terminal."""
        resp = requests.post(
            f"{OLLAMA_BASE}/api/chat",
            json={
                "model": self.model,
                "messages": messages,
                "stream": True,
                "options": {"num_predict": 300},
                "think": False,
            },
            stream=True,
        )
        resp.raise_for_status()
        full = ""
        for line in resp.iter_lines():
            if line:
                data = json.loads(line)
                token = data.get("message", {}).get("content", "")
                if token:
                    print(token, end="", flush=True)
                    full += token
        print()
        return full


# ============================================================
# CHATBOT CON RAG
# ============================================================
class ChatBot:
    """ChatBot con RAG: retrieval + generation + historial + referencias."""

    def __init__(self, collection, llm, top_k=5):
        self.collection = collection
        self.llm = llm
        self.top_k = top_k
        self.history = []

    def reset(self):
        self.history.clear()

    def follow_up_query(self, question):
        prompt = HISTORY_TEMPLATE.format(
            chat_history="\n".join(self.history),
            question=question,
        )
        messages = [{"role": "user", "content": prompt}]
        return self.llm.completion(messages)

    def search(self, query, k=None):
        k = k or self.top_k
        query_embedding = ollama_embed_single(query)
        results = self.collection.query(
            query_embeddings=[query_embedding],
            n_results=k,
            include=["documents", "metadatas", "distances"],
        )
        docs = []
        for i in range(len(results["documents"][0])):
            docs.append({
                "text": results["documents"][0][i],
                "source": results["metadatas"][0][i]["source"],
                "doc_type": results["metadatas"][0][i]["doc_type"],
                "score": 1 - results["distances"][0][i],
            })
        return docs

    def __call__(self, question, stream=False):
        # Detectar si el usuario pidió referencias con /reference
        show_refs = "/reference" in question.lower()
        query = question.replace("/reference", "").replace("/Reference", "").strip()

        # Retrieval
        documents = self.search(query)

        # Prompt con contexto
        context = "\n\n".join(doc["text"] for doc in documents)
        prompt = QUESTION_TEMPLATE.format(context=context, question=query)

        # Construir mensajes con historial para que el LLM recuerde la conversación
        messages = [{"role": "system", "content": SYSTEM_PROMPT}]
        for h in self.history[-4:]:  # últimas 4 interacciones máximo
            messages.append({"role": "user", "content": h.split("\nRespuesta:")[0].replace("Pregunta: ", "")})
            messages.append({"role": "assistant", "content": h.split("\nRespuesta:")[1].strip() if "\nRespuesta:" in h else ""})
        messages.append({"role": "user", "content": prompt})

        # Generar respuesta
        if stream:
            answer = self.llm.completion_stream_print(messages)
        else:
            answer = self.llm.completion(messages)

        # Historial (guardar sin referencias)
        self.history.append(f"Pregunta: {query}\nRespuesta: {answer}")

        # Referencias solo si se pidió con /reference
        if show_refs:
            refs = []
            seen = set()
            for doc in documents:
                if doc["source"] not in seen:
                    seen.add(doc["source"])
                    refs.append(
                        f"- {doc['source']} ({doc['doc_type']}, score: {doc['score']:.3f})"
                    )
            ref_block = "\n\n**Referencias:**\n" + "\n".join(refs)
            if stream:
                print(ref_block)
            return answer + ref_block

        return answer


# ============================================================
# MODO TERMINAL (Qwen 3.5 4B - ligero)
# ============================================================
def run_terminal(collection):
    """Chat en terminal con Qwen 3.5 4B."""
    print("\n" + "=" * 60)
    print("  Mini RAG - Terminal Mode (qwen3.5:4b)")
    print("  Escribe tu pregunta o 'salir' para terminar")
    print("=" * 60 + "\n")

    llm = OllamaLLM(model="qwen3.5:4b")
    bot = ChatBot(collection, llm, top_k=TOP_K)

    while True:
        try:
            question = input("Tu > ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nHasta luego!")
            break

        if not question:
            continue
        if question.lower() in ("salir", "exit", "quit"):
            print("Hasta luego!")
            break

        print("\nAsistente > ", end="")
        bot(question, stream=True)
        print()


# ============================================================
# MODO GRADIO (Qwen 3.5 9B - completo)
# ============================================================
def run_gradio(collection):
    """Chat con Gradio y Qwen 3.5 9B."""
    import gradio as gr

    print("\n[3/3] Lanzando interfaz Gradio con qwen3.5:9b...")

    llm = OllamaLLM(model="qwen3.5:9b")
    bot = ChatBot(collection, llm, top_k=TOP_K)

    with gr.Blocks(title="Mini RAG - Model Registry") as demo:
        gr.Markdown("# Mini RAG - Model Registry + STAMM Paper")
        gr.Markdown(
            "Pregunta sobre los modelos de ML registrados (RF, SVM, LSTM, GBM, CART) "
            "o sobre el paper STAMM."
        )

        chatbot_ui = gr.Chatbot(height=450)
        msg = gr.Textbox(
            label="Tu pregunta",
            placeholder="Ej: ¿Qué es STAMM? ¿Qué modelos usan Random Forest?",
        )
        clear = gr.Button("Limpiar conversación")

        def respond(question, chat_history):
            if not question.strip():
                return "", chat_history
            bot_message = bot(question)
            chat_history.append({"role": "user", "content": question})
            chat_history.append({"role": "assistant", "content": bot_message})
            return "", chat_history

        def reset_chat():
            bot.reset()
            return []

        msg.submit(respond, [msg, chatbot_ui], [msg, chatbot_ui])
        clear.click(reset_chat, None, chatbot_ui, queue=False)

    demo.launch(inbrowser=True, share=True)


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    terminal_mode = "--terminal" in sys.argv
    force_reingest = "--reingest" in sys.argv

    model_label = "qwen3.5:4b (terminal)" if terminal_mode else "qwen3.5:9b (Gradio)"

    print("=" * 60)
    print("  Mini RAG - Model Registry + STAMM Paper")
    print(f"  Modo: {model_label}")
    print("=" * 60)

    # Paso 1: Ingesta
    print("\n[1/3] Ingesta de YAMLs + PDFs...")
    client = chromadb.PersistentClient(path=CHROMA_DIR)

    if force_reingest:
        print("  Forzando re-ingesta...")
        collection = ingest()
    else:
        try:
            collection = client.get_collection("model_registry")
            count = collection.count()
            if count > 0:
                print(f"  ChromaDB ya tiene {count} chunks. Saltando ingesta.")
                print("  (usa --reingest para forzar re-ingesta)")
            else:
                collection = ingest()
        except Exception:
            collection = ingest()

    # Paso 2: Info
    print(f"\n[2/3] Configuración:")
    print(f"  Embeddings: {EMBED_MODEL}")
    print(f"  Top-K: {TOP_K}")

    # Paso 3: Lanzar
    if terminal_mode:
        run_terminal(collection)
    else:
        run_gradio(collection)
